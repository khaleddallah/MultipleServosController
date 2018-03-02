import openpyxl
import numpy as np

def exlHandler (file='excel.xlsx',sheet=0):
	wb=openpyxl.load_workbook(file)
	ws=wb.worksheets[sheet]
	
	data= np.tile(0,(ws.max_row-1,ws.max_column)).astype(float)
	#data = data.astype(float)
	
	waitTime= np.tile(0,(ws.max_row-1)).astype(float)
	#data = data.astype(float)

	targetTime= np.tile(0,(ws.max_row-1)).astype(float)
	
	

	for i in range (2,ws.max_row+1):
		for j in range (ws.max_column):
			if (j<ws.max_column-2):
				if (j<6):data[i-2][j]=-float(ws[i][j].value)+90
				else: data[i-2][j]=float(ws[i][j].value)+90
			elif (j==ws.max_column-2):
				waitTime[i-2]=ws[i][j].value
			elif (j==ws.max_column-1):
				targetTime[i-2] = ws[i][j].value
			else :
				print ("error")
				
				

	return(data,waitTime,targetTime)
		
	
if __name__== "__main__":
	print (exlHandler())
	
